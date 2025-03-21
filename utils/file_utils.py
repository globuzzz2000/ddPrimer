#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File handling utilities for ddPrimer pipeline.
"""

import os
import sys
import platform
import tkinter as tk
import pandas as pd
from tkinter import filedialog
from ..config import Config

# Optional imports for Excel formatting
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class FileUtils:
    """Handles file operations including UI dialogs and file parsing."""
    
    # Default to GUI mode unless explicitly detected as headless
    use_cli = False

    # Check for macOS and PyObjC availability
    is_macos = platform.system() == "Darwin"
    has_pyobjc = False
    
    if is_macos:
        try:
            import Foundation
            import AppKit
            has_pyobjc = True
        except ImportError:
            has_pyobjc = False

    # Detect headless environments explicitly with refined logic
    if platform.system() == "Linux":
        if not os.environ.get("DISPLAY", ""):
            use_cli = True
            reason = "No DISPLAY environment variable on Linux."
        else:
            reason = "DISPLAY variable found on Linux."
    elif platform.system() == "Windows":
        if not sys.stdout.isatty():
            use_cli = True
            reason = "Non-interactive terminal on Windows."
        else:
            reason = "Interactive terminal on Windows."
    elif platform.system() == "Darwin":
        # Force GUI mode by default on macOS unless explicitly headless
        use_cli = False
        reason = "Defaulting to GUI mode on macOS."
    else:
        reason = "Unknown platform, defaulting to GUI mode."

    # Initialize last directory with user's home directory
    _last_directory = None

    @classmethod
    def get_last_directory(cls):
        """
        Get the last directory used, loading from persistent storage if needed.
        
        Returns:
            str: Path to the last directory used
        """
        if cls._last_directory is None:
            # Try to load the last directory from a config file
            config_dir = os.path.join(os.path.expanduser("~"), ".ddPrimer")
            os.makedirs(config_dir, exist_ok=True)
            config_file = os.path.join(config_dir, "last_directory.txt")
            
            if os.path.exists(config_file):
                try:
                    with open(config_file, 'r') as f:
                        last_dir = f.read().strip()
                    if os.path.isdir(last_dir):
                        cls._last_directory = last_dir
                        Config.debug(f"Loaded last directory from config: {last_dir}")
                    else:
                        cls._last_directory = os.path.expanduser("~")
                except Exception as e:
                    Config.debug(f"Error reading last directory config: {e}")
                    cls._last_directory = os.path.expanduser("~")
            else:
                cls._last_directory = os.path.expanduser("~")
        
        return cls._last_directory
    
    @classmethod
    def save_last_directory(cls, directory):
        """
        Save the last directory to persistent storage.
        
        Args:
            directory (str): Path to save
        """
        if directory and os.path.isdir(directory):
            cls._last_directory = directory
            
            # Save to config file
            config_dir = os.path.join(os.path.expanduser("~"), ".ddPrimer")
            os.makedirs(config_dir, exist_ok=True)
            config_file = os.path.join(config_dir, "last_directory.txt")
            
            try:
                with open(config_file, 'w') as f:
                    f.write(directory)
                Config.debug(f"Saved last directory to config: {directory}")
            except Exception as e:
                Config.debug(f"Error saving last directory config: {e}")

    @staticmethod
    def _show_macos_file_dialog(prompt, filetypes, multiple=False):
        """
        Show a file dialog using tkinter (fallback for macOS).
        
        Args:
            prompt (str): Title for the dialog
            filetypes (list): List of file type tuples for the dialog
            multiple (bool): Allow multiple file selection
            
        Returns:
            str or list: Selected file path(s) or None if cancelled
        """
        try:
            root = tk.Tk()
            root.withdraw()

            # Redirect stderr to suppress warnings
            original_stderr = sys.stderr
            sys.stderr = open(os.devnull, "w")

            if multiple:
                file_paths = filedialog.askopenfilenames(title=prompt, filetypes=filetypes, parent=root)
            else:
                file_paths = filedialog.askopenfilename(title=prompt, filetypes=filetypes, parent=root)

            sys.stderr.close()
            sys.stderr = original_stderr
            root.destroy()

            if not file_paths:
                return None if not multiple else []

            return file_paths
        except Exception as e:
            Config.debug(f"Error in fallback tkinter dialog: {e}")
            return None if not multiple else []

    @classmethod
    def get_file(cls, prompt, filetypes):
        """
        Show a file dialog or prompt for a file path if CLI mode is enabled.
        
        Args:
            prompt (str): Text to display in the file dialog
            filetypes (list): List of file type tuples for the dialog
            
        Returns:
            str: Selected file path
        """
        # Get the last directory from our persistent storage
        last_directory = cls.get_last_directory()
        Config.debug(f"Starting with last directory: {last_directory}")
        
        if cls.use_cli:
            file_path = input(f"{prompt} (enter file path manually): ").strip()
            if not file_path:
                print(f"{prompt} - No file selected. Exiting.")
                sys.exit(1)
            return file_path
        
        # Fall back to tkinter implementation
        try:
            # Ensure we're starting fresh - kill any existing tk instances
            try:
                if hasattr(tk, '_default_root') and tk._default_root:
                    tk._default_root.destroy()
                    tk._default_root = None
            except:
                pass  # Ignore if no existing root
            
            # Create completely new tk root
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            
            # Let the window manager position the dialog
            # This is more reliable than trying to position it ourselves
            
            # On macOS, we must use a simpler filetype pattern
            valid_filetypes = [("All Files", "*")]
            
            # Add more specific filetypes if not on macOS or if they're simple enough
            if not cls.is_macos:
                for desc, ext in filetypes:
                    if ext and isinstance(ext, str) and ext.strip() and ext != "*":
                        valid_filetypes.insert(0, (desc, ext))
            
            # Make sure dialog doesn't output errors to terminal
            original_stderr = sys.stderr
            sys.stderr = open(os.devnull, "w")
            
            # Log the initial directory we're using
            Config.debug(f"Opening dialog with initial directory: {last_directory}")
            
            try:
                # Show the dialog
                file_path = filedialog.askopenfilename(
                    title=prompt,
                    filetypes=valid_filetypes,
                    initialdir=last_directory
                )
            finally:
                # Restore stderr
                sys.stderr.close()
                sys.stderr = original_stderr
            
            # Force proper cleanup
            root.destroy()
            
            if not file_path:
                print(f"No file was selected in the dialog. Exiting.")
                sys.exit(1)
            
            # Update the last directory for next time
            if os.path.isfile(file_path):
                new_directory = os.path.dirname(file_path)
                cls.save_last_directory(new_directory)
                Config.debug(f"Updated last directory to: {new_directory}")
            
            print(f"File selected: {file_path}")
            return file_path
            
        except Exception as e:
            print(f"GUI initialization failed ({e}), falling back to CLI mode.")
            cls.use_cli = True
            return cls.get_file(prompt, filetypes)

    @classmethod
    def get_files(cls, prompt, filetypes):
        """
        Show a file dialog for multiple files or prompt if CLI mode is enabled.
        
        Args:
            prompt (str): Text to display in the file dialog
            filetypes (list): List of file type tuples for the dialog
            
        Returns:
            list: List of selected file paths
        """
        # Get the last directory from our persistent storage
        last_directory = cls.get_last_directory()
        Config.debug(f"Starting with last directory: {last_directory}")
            
        if cls.use_cli:
            paths = input(f"{prompt} (enter file paths separated by spaces): ").strip()
            file_paths = paths.split() if paths else []
            if not file_paths:
                print(f"{prompt} - No files selected. Exiting.")
                sys.exit(1)
            return file_paths
        
        # Fall back to tkinter implementation
        try:
            # Ensure we're starting fresh - kill any existing tk instances
            try:
                if hasattr(tk, '_default_root') and tk._default_root:
                    tk._default_root.destroy()
                    tk._default_root = None
            except:
                pass  # Ignore if no existing root
            
            # Create completely new tk root
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            
            # Let the window manager position the dialog
            # This is more reliable than trying to position it ourselves
            
            # On macOS, we must use a simpler filetype pattern
            valid_filetypes = [("All Files", "*")]
            
            # Add more specific filetypes if not on macOS or if they're simple enough
            if not cls.is_macos:
                for desc, ext in filetypes:
                    if ext and isinstance(ext, str) and ext.strip() and ext != "*":
                        valid_filetypes.insert(0, (desc, ext))
            
            # Make sure dialog doesn't output errors to terminal
            original_stderr = sys.stderr
            sys.stderr = open(os.devnull, "w")
            
            # Log the initial directory we're using
            Config.debug(f"Opening dialog with initial directory: {last_directory}")
            
            try:
                # Show the dialog
                file_paths = filedialog.askopenfilenames(
                    title=prompt,
                    filetypes=valid_filetypes,
                    initialdir=last_directory
                )
            finally:
                # Restore stderr
                sys.stderr.close()
                sys.stderr = original_stderr
            
            # Force proper cleanup
            root.destroy()
            
            if not file_paths:
                print(f"No files were selected in the dialog. Exiting.")
                sys.exit(1)
            
            # Update the last directory for next time
            if file_paths and os.path.isfile(file_paths[0]):
                new_directory = os.path.dirname(file_paths[0])
                cls.save_last_directory(new_directory)
                Config.debug(f"Updated last directory to: {new_directory}")
            
            print(f"Selected {len(file_paths)} file(s)")
            for path in file_paths:
                print(f" - {path}")
                
            return file_paths
            
        except Exception as e:
            print(f"GUI initialization failed ({e}), falling back to CLI mode.")
            cls.use_cli = True
            return cls.get_files(prompt, filetypes)
    
    @staticmethod
    def load_fasta(filepath):
        """
        Load sequences from a FASTA file into a dict: {header_without_gt: sequence}.
        Optimized for memory efficiency.
        
        Args:
            filepath (str): Path to the FASTA file
            
        Returns:
            dict: Dictionary mapping sequence headers to sequences
        """
        sequences = {}
        name = None
        seq_chunks = []
        
        Config.debug(f"Loading FASTA file: {filepath}")
        try:
            with open(filepath, 'r') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    if line.startswith(">"):
                        if name:
                            sequences[name] = "".join(seq_chunks).upper()
                        name = line[1:].split()[0]
                        seq_chunks = []
                    else:
                        seq_chunks.append(line)
                if name:
                    sequences[name] = "".join(seq_chunks).upper()
        except Exception as e:
            print(f"Error reading FASTA file {os.path.abspath(filepath)}: {e}")
            sys.exit(1)
        
        Config.debug(f"Loaded {len(sequences)} sequences from FASTA file")
        return sequences
    
    @staticmethod
    def save_formatted_excel(df, output_file, logger=None):
        """
        Save DataFrame to Excel with specific formatting.
        If openpyxl is not available, falls back to standard pandas Excel export.
        
        Args:
            df (pandas.DataFrame): DataFrame with primer results
            output_file (str): Path to save the formatted Excel file
            logger (logging.Logger, optional): Logger for debug messages
            
        Returns:
            str: Path to the saved Excel file
        """
        
        # When openpyxl is available, apply full formatting
        try:
            # Log actions if logger is provided
            if logger:
                logger.debug(f"Saving formatted Excel file to: {output_file}")
            
            # First, save with pandas to get a basic Excel file
            df.to_excel(output_file, index=False, engine='openpyxl')
            
            # Now open the file for formatting
            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            
            # Get the number of rows and columns
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Freeze panes - freeze first column and first two rows
            worksheet.freeze_panes = 'B3'
            
            # Create styles
            header_font = Font(bold=True)
            sequence_fill = PatternFill(
                start_color='D9D9D9',  # Light gray
                end_color='D9D9D9',
                fill_type='solid'
            )
            centered_alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply column width
            for col_num in range(1, max_col + 1):
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = 10.83
            
            # Apply row height
            for row_num in range(1, max_row + 1):
                worksheet.row_dimensions[row_num].height = 16
            
            # Identify column groupings for merging headers
            column_groups = {
                "Gene": None,  # No merging for Gene
                "Primer F": ["Primer F", "Tm F", "Penalty F", "Primer F dG", "Primer F BLAST"],
                "Primer R": ["Primer R", "Tm R", "Penalty R", "Primer R dG", "Primer R BLAST", "Pair Penalty"],
            }
            
            # Add Probe group if present
            if "Probe" in df.columns:
                column_groups["Probe"] = ["Probe", "Probe Tm", "Probe Penalty", "Probe dG", "Probe BLAST"]
            
            # Add Amplicon group
            column_groups["Amplicon"] = ["Amplicon", "Length", "Amplicon GC%", "Amplicon dG", 
                                        "Chromosome", "Location"]
            
            # Prepare for header merging - get column indices
            header_merges = []
            gene_col = None
            
            # First, find the Gene column index
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name == "Gene":
                    gene_col = col_idx
                    break
            
            # If Gene column wasn't found, use the first column
            if gene_col is None:
                gene_col = 1
            
            # Now find column groups for merging
            for main_header, subheaders in column_groups.items():
                if subheaders:
                    indices = [idx + 1 for idx, name in enumerate(df.columns) if name in subheaders]
                    if indices:
                        start_idx = min(indices)
                        end_idx = max(indices)
                        header_merges.append((start_idx, end_idx, main_header))
            
            # Insert a new row for main headers
            worksheet.insert_rows(1)
            max_row += 1
            
            # Add main headers BEFORE merging
            for start_col, end_col, header_text in header_merges:
                if start_col > 0 and end_col > 0 and start_col <= max_col and end_col <= max_col:
                    # Add header text first
                    cell = worksheet.cell(row=1, column=start_col)
                    cell.value = header_text
                    cell.font = header_font
                    cell.alignment = centered_alignment
                    
                    # Then merge cells
                    merge_range = f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1"
                    worksheet.merge_cells(merge_range)
            
            # Handle Gene header separately (merges vertically)
            if gene_col is not None and gene_col <= max_col:
                # Set value first
                gene_cell = worksheet.cell(row=1, column=gene_col)
                gene_cell.value = "Gene"  # Use explicit "Gene" text
                gene_cell.font = header_font
                gene_cell.alignment = centered_alignment
                
                # Then merge
                gene_merge_range = f"{get_column_letter(gene_col)}1:{get_column_letter(gene_col)}2"
                worksheet.merge_cells(gene_merge_range)
            
            # Apply bold to all header cells (row 2)
            for col_num in range(1, max_col + 1):
                cell = worksheet.cell(row=2, column=col_num)
                cell.font = header_font
                cell.alignment = centered_alignment
                cell.border = None
            
            # Special formatting for sequence columns
            sequence_cols = []
            
            # Identify sequence columns
            for col_idx, col_name in enumerate(df.columns, 1):
                # Apply sequence fill to Primer F, Primer R, Probe, and Amplicon columns
                if col_name in ["Primer F", "Primer R", "Probe", "Amplicon"]:
                    sequence_cols.append(col_idx)
            
            # Apply special formatting to sequence cells
            for row_num in range(3, max_row + 1):  # Start from first data row
                for col_idx in sequence_cols:
                    if col_idx <= max_col:  # Ensure column exists
                        cell = worksheet.cell(row=row_num, column=col_idx)
                        cell.fill = sequence_fill

            # Center-align all non-header, non-merged cells
            for row_num in range(3, max_row + 1):
                for col_num in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell.alignment = centered_alignment
            
            # Save the workbook
            workbook.save(output_file)
            
            return output_file
            
        except Exception as e:
            # If formatting fails, fall back to standard save
            if logger:
                logger.error(f"Error applying Excel formatting: {e}")
                logger.warning("Falling back to standard Excel export")
            
            # Basic fallback save
            df.to_excel(output_file, index=False)
            
            if logger:
                logger.info(f"Excel file saved to: {output_file} (without formatting)")
            
            return output_file