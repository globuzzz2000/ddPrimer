#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File I/O module for ddPrimer pipeline.

Contains functionality for:
1. Cross-platform file selection with GUI and CLI support
2. FASTA file reading and writing operations
3. Excel file formatting with comprehensive styling
4. Sequence table loading from CSV and Excel formats

This module provides consistent interfaces for file operations including
file selection, reading/writing, and format conversions across different
platforms and environments.
"""

import os
import sys
import platform
import logging
import pandas as pd
import contextlib
from ..config import FileSelectionError, FileFormatError, FileError

# Set up module logger
logger = logging.getLogger(__name__)

# Optional import for Excel formatting
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Optional import for wxPython file dialogs
try:
    import wx
    HAS_WX = True
except ImportError:
    HAS_WX = False


@contextlib.contextmanager
def _silence_cocoa_stderr():
    """
    Temporarily redirect C-level stderr to /dev/null (macOS IMK chatter).
    
    Yields:
        None - Context manager for silencing stderr on macOS
    """
    if platform.system() != "Darwin":
        yield
        return
        
    import os, sys
    old_fd = os.dup(2)
    devnull = os.open(os.devnull, os.O_WRONLY)
    os.dup2(devnull, 2)
    os.close(devnull)
    try:
        yield
    finally:
        os.dup2(old_fd, 2)
        os.close(old_fd)


class FileIO:
    """
    Handles all file I/O operations with consistent interfaces.
    
    This class provides cross-platform file operations including GUI and CLI
    file selection, FASTA file processing, Excel formatting, and sequence
    table loading. It automatically detects the environment and chooses
    appropriate methods for file operations.
    
    Example:
        >>> sequences = FileIO.load_fasta("genome.fasta")
        >>> file_path = FileIO.select_file("Select genome file", [("FASTA", "*.fasta")])
        >>> FileIO.save_results(df, "output_dir", "input.fasta")
    """
    
    # Default to GUI mode unless explicitly detected as headless
    use_cli = False

    # Check for macOS and PyObjC availability
    is_macos = platform.system() == "Darwin"
    has_pyobjc = False
    
    if is_macos:
        try:
            import Foundation
            import AppKit
            has_pyobjc = True
        except ImportError:
            has_pyobjc = False

    # Detect headless environments
    if platform.system() == "Linux":
        if not os.environ.get("DISPLAY", ""):
            use_cli = True
    elif platform.system() == "Windows":
        if not sys.stdout.isatty():
            use_cli = True

    # Initialize last directory
    _last_directory = None
    
    # Shared wxPython app instance
    _wx_app = None
    
    @classmethod
    def initialize_wx_app(cls):
        """
        Initialize the wxPython app if it doesn't exist yet.
        
        Raises:
            FileSelectionError: If wxPython initialization fails
        """
        if HAS_WX and cls._wx_app is None:
            try:
                with _silence_cocoa_stderr():
                    cls._wx_app = wx.App(False)
                    logger.debug("wxPython app initialized successfully")
            except Exception as e:
                error_msg = f"Failed to initialize wxPython app: {str(e)}"
                logger.error(error_msg)
                logger.debug(f"Error details: {str(e)}", exc_info=True)
                raise FileSelectionError(error_msg) from e
    
    @classmethod
    def hide_app(cls):
        """Hide the wxPython app without destroying it."""
        if cls.is_macos and cls.has_pyobjc and cls._wx_app is not None:
            try:
                import AppKit
                NSApplication = AppKit.NSApplication.sharedApplication()
                NSApplication.setActivationPolicy_(AppKit.NSApplicationActivationPolicyAccessory)
                logger.debug("App hidden from macOS dock")
            except Exception as e:
                logger.warning(f"Error hiding app from dock: {str(e)}")
                logger.debug(f"Error details: {str(e)}", exc_info=True)

    @classmethod
    def mark_selection_complete(cls):
        """Mark that all file selections are complete and hide the app."""
        logger.debug("File selection process marked as complete")
        cls.hide_app()

    @classmethod
    def get_last_directory(cls):
        """
        Get the last directory used, loading from unified config storage.
        
        Returns:
            Path to the last directory used, or home directory as fallback
        """
        if cls._last_directory is None:
            try:
                from ..config import Config
                last_dir = Config.load_user_setting("last_directory", None)
                
                if last_dir and os.path.isdir(last_dir):
                    cls._last_directory = last_dir
                    logger.debug(f"Loaded last directory from config: {last_dir}")
                else:
                    cls._last_directory = os.path.expanduser("~")
                    logger.debug(f"Using home directory: {cls._last_directory}")
                    
            except Exception as e:
                logger.warning(f"Error loading last directory: {str(e)}")
                logger.debug(f"Error details: {str(e)}", exc_info=True)
                cls._last_directory = os.path.expanduser("~")
        
        return cls._last_directory
    
    @classmethod
    def save_last_directory(cls, directory):
        """
        Save the last directory to unified config storage.
        
        Args:
            directory: Path to save as last used directory
        """
        if not directory or not isinstance(directory, str) or not os.path.isdir(directory):
            return
            
        cls._last_directory = directory
        
        try:
            from ..config import Config
            Config.save_user_setting("last_directory", directory)
            logger.debug(f"Saved last directory to config: {directory}")
        except Exception as e:
            logger.warning(f"Error saving last directory: {str(e)}")
            logger.debug(f"Error details: {str(e)}", exc_info=True)

    @classmethod
    def normalize_filetypes(cls, filetypes):
        """
        Normalize file types for different platforms.
        
        Args:
            filetypes: List of (description, extension) tuples
            
        Returns:
            Normalized file types list for the current platform
        """
        # Define "All Files" for different platforms
        if cls.is_macos or platform.system() == "Linux":
            all_files = ("All Files", "*")
        else:  # Windows
            all_files = ("All Files", "*.*")
            
        normalized = [all_files]
        
        if filetypes:
            for desc, ext in filetypes:
                if ext == "*" or ext == "*.*":
                    continue
                
                if cls.is_macos:
                    clean_ext = ext.lstrip("*.")
                    if "." in clean_ext:
                        base_ext = clean_ext.split(".")[-1]
                        if base_ext:
                            normalized.insert(0, (desc, base_ext))
                    elif clean_ext:
                        normalized.insert(0, (desc, clean_ext))
                else:
                    if not ext.startswith("*.") and ext != "*":
                        clean_ext = f"*.{ext.lstrip('.')}"
                    else:
                        clean_ext = ext
                    normalized.insert(0, (desc, clean_ext))
                    
        return normalized
        
    @classmethod
    def select_file(cls, prompt, filetypes):
        """
        Show a file dialog or prompt for a file path if CLI mode is enabled.
        
        Args:
            prompt: Text to display in the file dialog
            filetypes: List of file type tuples for the dialog
            
        Returns:
            Selected file path
            
        Raises:
            FileSelectionError: If file selection fails or user cancels
        """
        last_directory = cls.get_last_directory()

        if cls.use_cli:
            try:
                file_path = input(f"{prompt}: ").strip()
                if not file_path:
                    error_msg = "No file path provided in CLI mode"
                    logger.error(error_msg)
                    raise FileSelectionError(error_msg)
                    
                if not os.path.exists(file_path):
                    error_msg = f"File not found: {file_path}"
                    logger.error(error_msg)
                    raise FileSelectionError(error_msg)
                
                return file_path
            except KeyboardInterrupt:
                error_msg = "File selection canceled by user"
                logger.error(error_msg)
                raise FileSelectionError(error_msg)
            except Exception as e:
                error_msg = f"CLI file selection failed: {str(e)}"
                logger.error(error_msg)
                logger.debug(f"Error details: {str(e)}", exc_info=True)
                raise FileSelectionError(error_msg) from e

        # wxPython implementation
        if not cls.use_cli and HAS_WX:
            try:
                cls.initialize_wx_app()
                
                valid_filetypes = cls.normalize_filetypes(filetypes)
                specific_types = [ft for ft in valid_filetypes if not ft[0].startswith("All Files")]
                wildcard = "|".join([f"{desc} ({ext})|{ext}" for desc, ext in specific_types]) or "All Files (*.*)|*.*"

                with _silence_cocoa_stderr():
                    dlg = wx.FileDialog(
                        None,
                        prompt,
                        wildcard=wildcard,
                        style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST
                    )
                    dlg.SetDirectory(last_directory)
                    dlg.SetFilterIndex(0)

                    if dlg.ShowModal() == wx.ID_OK:
                        file_path = dlg.GetPath()
                    else:
                        file_path = ""
                    dlg.Destroy()

                if not file_path:
                    error_msg = "No file was selected in the dialog"
                    logger.error(error_msg)
                    raise FileSelectionError(error_msg)

                cls.save_last_directory(os.path.dirname(file_path))
                return file_path

            except FileSelectionError:
                raise
            except Exception as e:
                logger.warning(f"wxPython file selection failed: {str(e)}")
                logger.debug(f"Error details: {str(e)}", exc_info=True)
                cls.use_cli = True
        
        # Fallback to CLI
        logger.warning("Falling back to CLI mode for file selection")
        return cls.select_file(prompt, filetypes)
        
    @staticmethod
    def load_fasta(filepath):
        """
        Load sequences from a FASTA file into a dictionary.
        
        Args:
            filepath: Path to the FASTA file
            
        Returns:
            Dictionary mapping sequence headers to sequences
            
        Raises:
            FileError: If the FASTA file doesn't exist
            FileFormatError: If there's an error parsing the FASTA file
        """
        sequences = {}
        name = None
        seq_chunks = []
        
        if not os.path.exists(filepath):
            error_msg = f"FASTA file not found: {filepath}"
            logger.error(error_msg)
            raise FileError(error_msg)
            
        try:
            with open(filepath, 'r') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    if line.startswith(">"):
                        if name:
                            sequences[name] = "".join(seq_chunks).upper()
                        name = line[1:].split()[0]
                        seq_chunks = []
                    else:
                        seq_chunks.append(line)
                        
                # Handle the last sequence
                if name:
                    sequences[name] = "".join(seq_chunks).upper()
                    
        except (OSError, IOError) as e:
            error_msg = f"Error reading FASTA file {os.path.abspath(filepath)}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileError(error_msg) from e
        except Exception as e:
            error_msg = f"Error parsing FASTA file {os.path.abspath(filepath)}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileFormatError(error_msg) from e
        
        logger.debug(f"Successfully loaded {len(sequences)} sequences from FASTA file")
        return sequences
    
    @staticmethod
    def save_fasta(sequences, filepath):
        """
        Save sequences to a FASTA file.
        
        Args:
            sequences: Dictionary of sequence ID to sequence
            filepath: Path to save the FASTA file
            
        Raises:
            FileFormatError: If there's an error writing the FASTA file
        """
        if not sequences:
            logger.warning("No sequences provided to save_fasta")
            
        try:
            with open(filepath, 'w') as f:
                for seq_id, sequence in sequences.items():
                    if not seq_id or not sequence:
                        logger.warning(f"Skipping invalid sequence: id='{seq_id}', seq_len={len(sequence) if sequence else 0}")
                        continue
                        
                    f.write(f">{seq_id}\n")
                    # Write sequence in chunks of 60 characters for readability
                    for i in range(0, len(sequence), 60):
                        f.write(f"{sequence[i:i+60]}\n")
                        
        except (OSError, IOError) as e:
            error_msg = f"Error writing FASTA file {filepath}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileFormatError(error_msg) from e
        except Exception as e:
            error_msg = f"Unexpected error writing FASTA file {filepath}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileFormatError(error_msg) from e
            
        logger.debug(f"Successfully saved FASTA file: {filepath}")
    
    @staticmethod
    def format_excel(df, output_file):
        """
        Save DataFrame to Excel with comprehensive formatting.
        
        Args:
            df: DataFrame with primer results
            output_file: Path to save the formatted Excel file
            
        Returns:
            Path to the saved Excel file
            
        Raises:
            FileFormatError: If Excel file cannot be created
        """
        try:
            if not HAS_OPENPYXL:
                raise ImportError("openpyxl is not available")
                
            # First, save with pandas to get a basic Excel file
            df.to_excel(output_file, index=False, engine='openpyxl')
            
            # Now open the file for formatting
            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Create a new row for our custom headers
            worksheet.insert_rows(1)
            max_row += 1
            
            # Create styles
            header_font = Font(bold=True)
            sequence_fill = PatternFill(
                start_color='D9D9D9',
                end_color='D9D9D9',
                fill_type='solid'
            )
            centered_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            column_map = {}
            header_texts = []
            
            # Apply basic formatting
            for col_num in range(1, max_col + 1):
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = 15
                
                # Set header row formatting
                cell2 = worksheet.cell(row=2, column=col_num)
                cell2.font = header_font
                cell2.alignment = centered_alignment
                
                # Handle Gene column
                if cell2.value == "Gene":
                    cell1 = worksheet.cell(row=1, column=col_num)
                    cell1.value = "Gene"
                    cell1.font = header_font
                    cell1.alignment = centered_alignment
                    worksheet.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=col_num)

                header_text = cell2.value
                header_texts.append(header_text)
                column_map[header_text] = col_num
                
                # Format data cells
                for row_num in range(3, max_row + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    
                    # Apply sequence fill to sequence columns
                    if header_text in ["Primer F", "Primer R", "Probe", "Amplicon"]:
                        cell.fill = sequence_fill
                    
                    # Handle "No suitable primers found" cells
                    if cell.value == "No suitable primers found":
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = centered_alignment
            
            # Freeze panes
            worksheet.freeze_panes = 'B3'
            
            # Group columns
            header_groups = {
                "Gene": [],
                "Forward Primer": ["Primer F", "Tm F", "Penalty F", "Primer F dG", "Primer F BLAST1", "Primer F BLAST2"],
                "Reverse Primer": ["Primer R", "Tm R", "Penalty R", "Primer R dG", "Primer R BLAST1", "Primer R BLAST2", "Pair Penalty"],
                "Probe": ["Probe", "Probe Tm", "Probe Penalty", "Probe dG", "Probe BLAST1", "Probe BLAST2"],
                "Amplicon": ["Amplicon", "Length", "Amplicon GC%", "Amplicon dG"],
                "Location": ["Chromosome", "Location"]
            }
            
            for group_name, headers in header_groups.items():
                existing_headers = [h for h in headers if h in header_texts]
                
                if not existing_headers:
                    continue
                    
                col_indices = [column_map[h] for h in existing_headers]
                
                if col_indices:
                    start_col = min(col_indices)
                    end_col = max(col_indices)
                    
                    group_cell = worksheet.cell(row=1, column=start_col)
                    group_cell.value = group_name
                    group_cell.font = header_font
                    group_cell.alignment = centered_alignment
                    
                    if start_col != end_col:
                        merge_range = f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1"
                        try:
                            worksheet.merge_cells(merge_range)
                        except Exception as e:
                            logger.warning(f"Could not merge range {merge_range}: {str(e)}")
            
            workbook.save(output_file)
            return output_file
                
        except ImportError:
            logger.warning("openpyxl not available, falling back to standard Excel export")
        except Exception as e:
            error_msg = f"Error applying Excel formatting: {str(e)}"
            logger.error(error_msg)
            logger.warning("Falling back to standard Excel export")
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            
        # Basic fallback save
        try:
            df.to_excel(output_file, index=False)
            logger.info(f"Excel file saved to: {output_file} (without formatting)")
            return output_file
        except Exception as e:
            error_msg = f"Failed to save Excel file {output_file}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileFormatError(error_msg) from e

    @staticmethod
    def load_sequences_from_table(file_path):
        """
        Load sequences from a CSV or Excel file with improved column detection.
        
        Args:
            file_path: Path to CSV or Excel file
            
        Returns:
            Dictionary of sequence ID to sequence
            
        Raises:
            FileError: If the file doesn't exist
            FileFormatError: If there's an error parsing the file
        """
        if not os.path.exists(file_path):
            error_msg = f"Table file not found: {file_path}"
            logger.error(error_msg)
            raise FileError(error_msg)
            
        try:
            # Use SequenceAnalyzer to analyze the file structure
            analysis = SequenceAnalyzer.analyze_file(file_path)
            
            if "error" in analysis:
                error_msg = f"Error analyzing file: {analysis['error']}"
                logger.error(error_msg)
                raise FileFormatError(error_msg)
                
            # Get recommended columns from analysis
            name_col, seq_col = SequenceAnalyzer.get_recommended_columns(analysis)
            
            # Load with pandas
            try:
                if file_path.endswith('.csv'):
                    df = pd.read_csv(file_path)
                elif file_path.endswith(('.xlsx', '.xls')):
                    df = pd.read_excel(file_path)
                else:
                    error_msg = f"Unsupported file format: {file_path}. Must be CSV or Excel."
                    logger.error(error_msg)
                    raise FileFormatError(error_msg)
            except Exception as e:
                error_msg = f"Error reading file {file_path}: {str(e)}"
                logger.error(error_msg)
                logger.debug(f"Error details: {str(e)}", exc_info=True)
                raise FileFormatError(error_msg) from e
            
            # Remove empty columns and rows
            df = df.dropna(axis=1, how='all')
            df = df.dropna(axis=0, how='all')
            
            sequences = {}
            
            if not seq_col:
                error_msg = "Could not identify a sequence column in the file"
                logger.error(error_msg)
                raise FileFormatError(error_msg)
                
            # Generate IDs if no name column
            if not name_col:
                for idx, row in df.iterrows():
                    sequence = str(row[seq_col]).strip().upper()
                    
                    if pd.isna(sequence) or not sequence:
                        continue
                    
                    seq_id = f"Seq_{idx+1}"
                    sequences[seq_id] = sequence
            else:
                # Use name and sequence columns
                for idx, row in df.iterrows():
                    seq_id = str(row[name_col]).strip()
                    sequence = str(row[seq_col]).strip().upper()
                    
                    if pd.isna(seq_id) or pd.isna(sequence) or not seq_id or not sequence:
                        continue
                    
                    # Handle duplicate IDs
                    original_seq_id = seq_id
                    if seq_id in sequences:
                        suffix = 1
                        while f"{seq_id}_{suffix}" in sequences:
                            suffix += 1
                        seq_id = f"{seq_id}_{suffix}"
                    
                    sequences[seq_id] = sequence
            
            logger.debug(f"Successfully loaded {len(sequences)} sequences from table file")
            return sequences
            
        except (FileFormatError, FileError):
            raise
        except Exception as e:
            error_msg = f"Error parsing sequence table {file_path}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileFormatError(error_msg) from e

    @staticmethod
    def save_results(df, output_dir, input_file, mode='standard'):
        """
        Save results to an Excel file with correct naming based on input files and mode.
        
        Args:
            df: DataFrame with primer results
            output_dir: Output directory
            input_file: Path to the input file (FASTA, CSV, etc.)
            mode: Pipeline mode ('standard' or 'direct')
            
        Returns:
            Path to the output file
            
        Raises:
            FileFormatError: If there's an error saving the results
        """
        # Make sure output directory exists
        try:
            os.makedirs(output_dir, exist_ok=True)
        except OSError as e:
            error_msg = f"Failed to create output directory {output_dir}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileFormatError(error_msg) from e
        
        # Set output filename
        basename = os.path.basename(input_file)
        root, _ = os.path.splitext(basename)
        output_filename = f"Primers_{root}.xlsx"
        output_file = os.path.join(output_dir, output_filename)
        
        # Define columns based on mode
        if mode == 'direct':
            # Direct mode excludes location columns (NOT YET IMPLEMENTED)
            columns = [
                "Gene", "Primer F", "Tm F", "Penalty F", "Primer F dG", "Primer F BLAST1", "Primer F BLAST2",
                "Primer R", "Tm R", "Penalty R", "Primer R dG", "Primer R BLAST1", "Primer R BLAST2",
                "Pair Penalty", "Amplicon", "Length", "Amplicon GC%", "Amplicon dG"
            ]
        else:
            # Standard mode includes location columns
            columns = [
                "Gene", "Primer F", "Tm F", "Penalty F", "Primer F dG", "Primer F BLAST1", "Primer F BLAST2",
                "Primer R", "Tm R", "Penalty R", "Primer R dG", "Primer R BLAST1", "Primer R BLAST2",
                "Pair Penalty", "Amplicon", "Length", "Amplicon GC%", "Amplicon dG",
                "Chromosome", "Location"
            ]
        
        # Add probe columns if present
        if "Probe" in df.columns:
            probe_cols = [
                "Probe", "Probe Tm", "Probe Penalty", "Probe dG", 
                "Probe BLAST1", "Probe BLAST2"
            ]
            # Insert probe columns after primer columns
            idx = columns.index("Pair Penalty") + 1
            for col in reversed(probe_cols):
                if col in df.columns:
                    columns.insert(idx, col)
        
        # Ensure all columns exist in DataFrame
        columns = [col for col in columns if col in df.columns]
        
        # Reorder the columns
        df = df[columns]
        
        # Save with formatting
        try:
            output_path = FileIO.format_excel(df, output_file)
            return output_path
        except Exception as e:
            error_msg = f"Error saving Excel file: {str(e)}"
            logger.error(error_msg)
            logger.warning("Falling back to basic Excel export")
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            
            try:
                df.to_excel(output_file, index=False)
                logger.info(f"\nResults saved to: {output_file} (without formatting)")
                return output_file
            except Exception as ex:
                error_msg = f"Failed to save results to {output_file}: {str(ex)}"
                logger.error(error_msg)
                logger.debug(f"Error details: {str(ex)}", exc_info=True)
                raise FileFormatError(error_msg) from ex