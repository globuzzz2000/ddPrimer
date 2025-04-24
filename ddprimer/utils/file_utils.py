#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fixed FileUtils class for ddPrimer pipeline.
This addresses the file selection crash on macOS after selecting the first file.
"""

import os
import sys
import platform
import tkinter as tk
import pandas as pd
from tkinter import filedialog
from ..config import Config

# Optional imports for Excel formatting
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class FileUtils:
    """Handles file operations including UI dialogs and file parsing."""
    
    # Default to GUI mode unless explicitly detected as headless
    use_cli = False

    # Check for macOS and PyObjC availability
    is_macos = platform.system() == "Darwin"
    has_pyobjc = False
    
    if is_macos:
        try:
            import Foundation
            import AppKit
            has_pyobjc = True
        except ImportError:
            has_pyobjc = False

    # Detect headless environments explicitly with refined logic
    if platform.system() == "Linux":
        if not os.environ.get("DISPLAY", ""):
            use_cli = True
            reason = "No DISPLAY environment variable on Linux."
        else:
            reason = "DISPLAY variable found on Linux."
    elif platform.system() == "Windows":
        if not sys.stdout.isatty():
            use_cli = True
            reason = "Non-interactive terminal on Windows."
        else:
            reason = "Interactive terminal on Windows."
    elif platform.system() == "Darwin":
        # Force GUI mode by default on macOS unless explicitly headless
        use_cli = False
        reason = "Defaulting to GUI mode on macOS."
    else:
        reason = "Unknown platform, defaulting to GUI mode."

    # Initialize last directory with user's home directory
    _last_directory = None

    @classmethod
    def get_last_directory(cls):
        """
        Get the last directory used, loading from persistent storage if needed.
        
        Returns:
            str: Path to the last directory used
        """
        if cls._last_directory is None:
            # Try to load the last directory from a config file
            config_dir = os.path.join(os.path.expanduser("~"), ".ddPrimer")
            os.makedirs(config_dir, exist_ok=True)
            config_file = os.path.join(config_dir, "last_directory.txt")
            
            if os.path.exists(config_file):
                try:
                    with open(config_file, 'r') as f:
                        last_dir = f.read().strip()
                    if os.path.isdir(last_dir):
                        cls._last_directory = last_dir
                        Config.debug(f"Loaded last directory from config: {last_dir}")
                    else:
                        cls._last_directory = os.path.expanduser("~")
                except Exception as e:
                    Config.debug(f"Error reading last directory config: {e}")
                    cls._last_directory = os.path.expanduser("~")
            else:
                cls._last_directory = os.path.expanduser("~")
        
        return cls._last_directory
    
    @classmethod
    def save_last_directory(cls, directory):
        """
        Save the last directory to persistent storage.
        
        Args:
            directory (str): Path to save
        """
        if directory and os.path.isdir(directory):
            cls._last_directory = directory
            
            # Save to config file
            config_dir = os.path.join(os.path.expanduser("~"), ".ddPrimer")
            os.makedirs(config_dir, exist_ok=True)
            config_file = os.path.join(config_dir, "last_directory.txt")
            
            try:
                with open(config_file, 'w') as f:
                    f.write(directory)
                Config.debug(f"Saved last directory to config: {directory}")
            except Exception as e:
                Config.debug(f"Error saving last directory config: {e}")

    @classmethod
    def normalize_filetypes(cls, filetypes):
        """
        Normalize file types for different platforms.
        
        Args:
            filetypes (list): List of (description, extension) tuples
            
        Returns:
            list: Normalized file types list
        """
        # Define "All Files" for different platforms
        if cls.is_macos or platform.system() == "Linux":
            all_files = ("All Files", "*")
        else:  # Windows
            all_files = ("All Files", "*.*")
            
        # Start with all files as a default
        normalized = [all_files]
        
        # Add specific file types
        if filetypes:
            for desc, ext in filetypes:
                if ext == "*" or ext == "*.*":
                    continue  # Skip generic "all files" entries
                
                if cls.is_macos:
                    # macOS: extension without wildcards (e.g., "txt" not "*.txt")
                    clean_ext = ext.lstrip("*.")
                    
                    # Handle compound extensions for macOS
                    if "." in clean_ext:
                        # For compound extensions like vcf.gz, use just the last part
                        base_ext = clean_ext.split(".")[-1]
                        if base_ext:
                            normalized.insert(0, (desc, base_ext))
                    elif clean_ext:
                        normalized.insert(0, (desc, clean_ext))
                else:
                    # Windows/Linux: ensure wildcard format (e.g., "*.txt")
                    if not ext.startswith("*.") and ext != "*":
                        clean_ext = f"*.{ext.lstrip('.')}"
                    else:
                        clean_ext = ext
                    normalized.insert(0, (desc, clean_ext))
                    
        return normalized

    @classmethod
    def get_file(cls, prompt, filetypes):
        """
        Show a file dialog or prompt for a file path if CLI mode is enabled.
        
        Args:
            prompt (str): Text to display in the file dialog
            filetypes (list): List of file type tuples for the dialog
            
        Returns:
            str: Selected file path
        """
        # Get the last directory from our persistent storage
        last_directory = cls.get_last_directory()
        Config.debug(f"Starting with last directory: {last_directory}")
        
        if cls.use_cli:
            file_path = input(f"{prompt} (enter file path manually): ").strip()
            if not file_path:
                print(f"{prompt} - No file selected. Exiting.")
                sys.exit(1)
            return file_path
        
        # Fall back to tkinter implementation
        try:
            # Ensure we're starting fresh - kill any existing tk instances
            try:
                if hasattr(tk, '_default_root') and tk._default_root:
                    tk._default_root.destroy()
                    tk._default_root = None
            except:
                pass  # Ignore if no existing root
            
            # Create completely new tk root
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            
            # Process file types for the current platform
            valid_filetypes = cls.normalize_filetypes(filetypes)
            Config.debug(f"Using file types: {valid_filetypes}")
            
            # Make sure dialog doesn't output errors to terminal
            original_stderr = sys.stderr
            sys.stderr = open(os.devnull, "w")
            
            # Log the initial directory we're using
            Config.debug(f"Opening dialog with initial directory: {last_directory}")
            
            try:
                # Show the dialog
                file_path = filedialog.askopenfilename(
                    title=prompt,
                    filetypes=valid_filetypes,
                    initialdir=last_directory
                )
            finally:
                # Restore stderr
                sys.stderr.close()
                sys.stderr = original_stderr
            
            # Force proper cleanup
            root.destroy()
            
            if not file_path:
                print(f"No file was selected in the dialog. Exiting.")
                sys.exit(1)
            
            # Update the last directory for next time
            directory_path = os.path.dirname(file_path)
            cls.save_last_directory(directory_path)
            Config.debug(f"Updated last directory to: {directory_path}")
            
            print(f"Selected file: {file_path}")
            return file_path
            
        except Exception as e:
            print(f"GUI initialization failed ({e}), falling back to CLI mode.")
            cls.use_cli = True
            return cls.get_file(prompt, filetypes)

    @classmethod
    def get_files(cls, prompt, filetypes):
        """
        Show a file dialog for multiple files or prompt if CLI mode is enabled.
        
        Args:
            prompt (str): Text to display in the file dialog
            filetypes (list): List of file type tuples for the dialog
            
        Returns:
            list: List of selected file paths
        """
        # Get the last directory from our persistent storage
        last_directory = cls.get_last_directory()
        Config.debug(f"Starting with last directory: {last_directory}")
            
        if cls.use_cli:
            paths = input(f"{prompt} (enter file paths separated by spaces): ").strip()
            file_paths = paths.split() if paths else []
            if not file_paths:
                print(f"{prompt} - No files selected. Exiting.")
                sys.exit(1)
            return file_paths
        
        # Fall back to tkinter implementation
        try:
            # Ensure we're starting fresh - kill any existing tk instances
            try:
                if hasattr(tk, '_default_root') and tk._default_root:
                    tk._default_root.destroy()
                    tk._default_root = None
            except:
                pass  # Ignore if no existing root
            
            # Create completely new tk root
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            
            # Process file types for the current platform
            valid_filetypes = cls.normalize_filetypes(filetypes)
            Config.debug(f"Using file types: {valid_filetypes}")
            
            # Make sure dialog doesn't output errors to terminal
            original_stderr = sys.stderr
            sys.stderr = open(os.devnull, "w")
            
            # Log the initial directory we're using
            Config.debug(f"Opening dialog with initial directory: {last_directory}")
            
            try:
                # Show the dialog
                file_paths = filedialog.askopenfilenames(
                    title=prompt,
                    filetypes=valid_filetypes,
                    initialdir=last_directory
                )
            finally:
                # Restore stderr
                sys.stderr.close()
                sys.stderr = original_stderr
            
            # Force proper cleanup
            root.destroy()
            
            if not file_paths:
                print(f"No files were selected in the dialog. Exiting.")
                sys.exit(1)
            
            # Update the last directory for next time
            if file_paths and os.path.isfile(file_paths[0]):
                new_directory = os.path.dirname(file_paths[0])
                cls.save_last_directory(new_directory)
                Config.debug(f"Updated last directory to: {new_directory}")
            
            print(f"Selected {len(file_paths)} file(s)")
            for path in file_paths:
                print(f" - {path}")
                
            return file_paths
            
        except Exception as e:
            print(f"GUI initialization failed ({e}), falling back to CLI mode.")
            cls.use_cli = True
            return cls.get_files(prompt, filetypes)
    
    @staticmethod
    def load_fasta(filepath):
        """
        Load sequences from a FASTA file into a dict: {header_without_gt: sequence}.
        Optimized for memory efficiency.
        
        Args:
            filepath (str): Path to the FASTA file
            
        Returns:
            dict: Dictionary mapping sequence headers to sequences
        """
        sequences = {}
        name = None
        seq_chunks = []
        
        Config.debug(f"Loading FASTA file: {filepath}")
        try:
            with open(filepath, 'r') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    if line.startswith(">"):
                        if name:
                            sequences[name] = "".join(seq_chunks).upper()
                        name = line[1:].split()[0]
                        seq_chunks = []
                    else:
                        seq_chunks.append(line)
                if name:
                    sequences[name] = "".join(seq_chunks).upper()
        except Exception as e:
            print(f"Error reading FASTA file {os.path.abspath(filepath)}: {e}")
            sys.exit(1)
        
        Config.debug(f"Loaded {len(sequences)} sequences from FASTA file")
        return sequences
    
    @staticmethod
    def save_formatted_excel(df, output_file, logger=None):
        """
        Save DataFrame to Excel with specific formatting.
        If openpyxl is not available, falls back to standard pandas Excel export.
        
        Args:
            df (pandas.DataFrame): DataFrame with primer results
            output_file (str): Path to save the formatted Excel file
            logger (logging.Logger, optional): Logger for debug messages
            
        Returns:
            str: Path to the saved Excel file
        """
        
        # When openpyxl is available, apply full formatting
        try:
            if not HAS_OPENPYXL:
                raise ImportError("openpyxl is not available")
                
            # Log actions if logger is provided
            if logger:
                logger.debug(f"Saving formatted Excel file to: {output_file}")
            
            # First, save with pandas to get a basic Excel file
            df.to_excel(output_file, index=False, engine='openpyxl')
            
            # Now open the file for formatting
            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            
            # Get the number of rows and columns
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Create a new row for our custom headers before doing anything else
            worksheet.insert_rows(1)
            max_row += 1
            
            # Create styles
            header_font = Font(bold=True)
            sequence_fill = PatternFill(
                start_color='D9D9D9',  # Light gray
                end_color='D9D9D9',
                fill_type='solid'
            )
            centered_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # Create a dictionary to map column names to their indices
            column_map = {}
            header_texts = []
            
            # First apply basic formatting to all cells before merging
            for col_num in range(1, max_col + 1):
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = 15  # Slightly wider columns
                
                # Set header row formatting in row 2 (original headers)
                cell2 = worksheet.cell(row=2, column=col_num)
                cell2.font = header_font
                cell2.alignment = centered_alignment
                

                # Merge "Gene" header if applicable
                if cell2.value == "Gene":
                    cell1 = worksheet.cell(row=1, column=col_num)
                    cell1.value = "Gene"
                    cell1.font = header_font
                    cell1.alignment = centered_alignment
                    cell1.border = openpyxl.styles.borders.Border()
                    worksheet.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=col_num)

                # Get the header text and store it
                header_text = cell2.value
                header_texts.append(header_text)
                column_map[header_text] = col_num
                
                # Format data cells
                for row_num in range(3, max_row + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    
                    # Apply sequence fill to Primer F, Primer R, Probe, and Amplicon columns
                    if header_text in ["Primer F", "Primer R", "Probe", "Amplicon"]:
                        cell.fill = sequence_fill
                    
                    # Special handling for "No suitable primers found" cells - left-align these
                    if cell.value == "No suitable primers found":
                        cell.alignment = left_alignment
                    else:
                        # Center alignment for all other cells
                        cell.alignment = centered_alignment
            
            # Freeze panes - freeze first column and first two rows
            worksheet.freeze_panes = 'B3'
            
            # Group the columns for our custom header row
            header_groups = {
                "Gene": [],
                "Forward Primer": ["Primer F", "Tm F", "Penalty F", "Primer F dG", "Primer F BLAST1", "Primer F BLAST2"],
                "Reverse Primer": ["Primer R", "Tm R", "Penalty R", "Primer R dG", "Primer R BLAST1", "Primer R BLAST2", "Pair Penalty"],
                "Probe": ["Probe", "Probe Tm", "Probe Penalty", "Probe dG", "Probe BLAST1", "Probe BLAST2"],
                "Amplicon": ["Amplicon", "Length", "Amplicon GC%", "Amplicon dG"],
                "Location": ["Chromosome", "Location", "Qry Chromosome", "Qry Location"]
            }
            
            # Add the group headers safely by first checking if any columns from each group exist
            for group_name, headers in header_groups.items():
                # Filter to only include headers that actually exist in our DataFrame
                existing_headers = [h for h in headers if h in header_texts]
                
                if not existing_headers:
                    if logger:
                        logger.debug(f"Skipping group '{group_name}' as none of its columns exist")
                    continue
                    
                # Find column indices for existing headers
                col_indices = [column_map[h] for h in existing_headers]
                
                if col_indices:
                    start_col = min(col_indices)
                    end_col = max(col_indices)
                    
                    # Add the group header
                    group_cell = worksheet.cell(row=1, column=start_col)
                    group_cell.value = group_name
                    group_cell.font = header_font
                    group_cell.alignment = centered_alignment
                    
                    # Merge if there's more than one column
                    if start_col != end_col:
                        merge_range = f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1"
                        try:
                            worksheet.merge_cells(merge_range)
                            if logger:
                                logger.debug(f"Merged cells for group '{group_name}' ({merge_range})")
                        except Exception as e:
                            # If merge fails, just leave as individual cells
                            if logger:
                                logger.debug(f"Could not merge range {merge_range}: {e}")
            
            # Save the workbook
            workbook.save(output_file)
            
            return output_file
                
        except Exception as e:
            # If formatting fails, fall back to standard save
            if logger:
                logger.error(f"Error applying Excel formatting: {e}")
                logger.warning("Falling back to standard Excel export")
            
            # Basic fallback save
            df.to_excel(output_file, index=False)
            
            if logger:
                logger.info(f"Excel file saved to: {output_file} (without formatting)")
            
            return output_file

        
    @staticmethod
    def load_sequences_from_table(file_path):
        """
        Load sequences from a CSV or Excel file.
        
        Args:
            file_path (str): Path to CSV or Excel file
            
        Returns:
            dict: Dictionary of sequence ID to sequence
        """
        import logging
        logger = logging.getLogger("ddPrimer")
        
        logger.debug(f"Loading sequences from table file: {file_path}")
        
        # Determine file type
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        elif file_path.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_path}. Must be CSV or Excel.")
        
        # Remove any empty columns
        df = df.dropna(axis=1, how='all')
        
        # If no headers, use the first two columns directly
        if df.columns.tolist() == [0, 1] or df.columns.tolist() == ['0', '1'] or df.columns.tolist() == [0, 1, 2]:
            seq_name_col = df.columns[0]
            seq_col = df.columns[1]
        else:
            # Try to find appropriate column names if headers exist
            columns = df.columns.tolist()
            seq_name_col = next((col for col in columns if col.lower() in 
                            ['sequence name', 'sequence_name', 'name', 'id', 'seq_id', 'sequence id']), None)
            seq_col = next((col for col in columns if col.lower() in 
                        ['sequence', 'seq', 'dna', 'dna_sequence', 'nucleotide']), None)
            
            # If we can't find matching columns by name, use first two non-empty columns
            if not seq_name_col or not seq_col:
                non_empty_cols = [col for col in columns if not df[col].isna().all()]
                if len(non_empty_cols) >= 2:
                    seq_name_col = non_empty_cols[0]
                    seq_col = non_empty_cols[1]
                else:
                    raise ValueError("Input file must have at least two columns with data")
        
        # Convert to dictionary
        sequences = {}
        for _, row in df.iterrows():
            seq_id = str(row[seq_name_col]).strip()
            sequence = str(row[seq_col]).strip().upper()
            
            # Skip empty rows
            if pd.isna(seq_id) or pd.isna(sequence) or not seq_id or not sequence:
                continue
                
            sequences[seq_id] = sequence
        
        logger.debug(f"Loaded {len(sequences)} sequences from table file")
        return sequences

    @staticmethod
    def get_sequences_file():
        """
        Prompt the user to select a CSV or Excel file with sequences.
        
        Returns:
            str: Path to the selected file
        """
        import logging
        logger = logging.getLogger("ddPrimer")
        
        logger.info("\n>>> Please select CSV or Excel file with sequences <<<")
        try:
            return FileUtils.get_file(
                "Select CSV or Excel file with sequences", 
                [("CSV Files", "*.csv"), ("Excel Files", "*.xlsx"), ("Excel Files", "*.xls"), ("All Files", "*")]
            )
        except Exception as e:
            logger.error(f"Error selecting direct sequence file: {e}")
            import traceback
            logger.debug(traceback.format_exc())
            raise

    @staticmethod
    def save_results(df, output_dir, input_file, mode='standard', second_fasta=None, logger=None):
        """
        Save results to an Excel file with correct naming based on input files and mode.
        
        Args:
            df: DataFrame with primer results
            output_dir: Output directory
            input_file: Path to the input file (FASTA, CSV, MAF, etc.)
            mode: Pipeline mode ('standard', 'direct', or 'alignment')
            second_fasta: Path to second FASTA file for alignment mode
            logger: Optional logger for debugging information
            
        Returns:
            str: Path to the output file
        """
        if logger is None:
            import logging
            logger = logging.getLogger("ddPrimer")
            
        logger.debug("\nSaving results...")
        
        # Make sure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Set output filename based on mode and input files
        if mode == 'alignment':
            # Alignment mode has specific naming conventions
            if input_file and input_file.endswith('.maf'):
                # Using pre-computed MAF file
                basename = os.path.basename(input_file)
                root, _ = os.path.splitext(basename)
                output_filename = f"Primers_{root}.xlsx"
            elif input_file and second_fasta:
                # Using FASTA alignment with two genomes - use both filenames
                ref_basename = os.path.basename(input_file)
                ref_root, _ = os.path.splitext(ref_basename)
                
                qry_basename = os.path.basename(second_fasta)
                qry_root, _ = os.path.splitext(qry_basename)
                
                output_filename = f"Primers_{ref_root}_vs_{qry_root}.xlsx"
            else:
                # Fallback for alignment mode when files aren't specified properly
                output_filename = f"Primers_Alignment.xlsx"
        elif mode == 'direct':
            # For direct mode, use the basename of the input file
            basename = os.path.basename(input_file)
            root, _ = os.path.splitext(basename)
            output_filename = f"Primers_{root}.xlsx"
        else:
            # Standard mode - use basename of input file
            basename = os.path.basename(input_file)
            root, _ = os.path.splitext(basename)
            output_filename = f"Primers_{root}.xlsx"
        
        # Combine output directory and filename
        output_file = os.path.join(output_dir, output_filename)
        logger.debug(f"Output file will be: {output_file}")
        
        # Define columns based on mode
        if mode == 'direct':
            # Direct mode excludes location columns
            columns = [
                "Gene", "Primer F", "Tm F", "Penalty F", "Primer F dG", "Primer F BLAST1", "Primer F BLAST2",
                "Primer R", "Tm R", "Penalty R", "Primer R dG", "Primer R BLAST1", "Primer R BLAST2",
                "Pair Penalty", "Amplicon", "Length", "Amplicon GC%", "Amplicon dG"
            ]
        else:
            # Standard and alignment modes include location columns
            columns = [
                "Gene", "Primer F", "Tm F", "Penalty F", "Primer F dG", "Primer F BLAST1", "Primer F BLAST2",
                "Primer R", "Tm R", "Penalty R", "Primer R dG", "Primer R BLAST1", "Primer R BLAST2",
                "Pair Penalty", "Amplicon", "Length", "Amplicon GC%", "Amplicon dG",
                "Chromosome", "Location"
            ]
            
            # Add cross-species specific columns for alignment mode
            if mode == 'alignment':
                cross_species_cols = ["Qry Chromosome", "Qry Location"]
                for col in cross_species_cols:
                    if col in df.columns and not df[col].isna().all():
                        columns.append(col)
        
        # Add probe columns if present
        if "Probe" in df.columns:
            probe_cols = [
                "Probe", "Probe Tm", "Probe Penalty", "Probe dG", 
                "Probe BLAST1", "Probe BLAST2"
            ]
            # Insert probe columns after primer columns
            idx = columns.index("Pair Penalty") + 1
            for col in reversed(probe_cols):
                if col in df.columns:
                    columns.insert(idx, col)
            
            logger.debug(f"Added probe columns to output: {', '.join([c for c in probe_cols if c in df.columns])}")
        
        # Ensure all columns in the list exist in the DataFrame
        columns = [col for col in columns if col in df.columns]
        
        # Reorder the columns
        df = df[columns]
        
        # Save with formatting
        try:
            output_path = FileUtils.save_formatted_excel(df, output_file, logger=logger)
            logger.info(f"Results saved to: {output_path}")
            return output_path
        except Exception as e:
            # Fallback if formatting fails
            logger.error(f"Error saving Excel file: {e}")
            logger.warning("Falling back to basic Excel export")
            
            try:
                df.to_excel(output_file, index=False)
                logger.info(f"Results saved to: {output_file} (without formatting)")
                return output_file
            except Exception as ex:
                logger.error(f"Failed to save results: {ex}")
                return None

    @staticmethod
    def setup_output_directories(args, reference_file=None, mode='standard'):
        """
        Set up output directory and temporary directory for the pipeline.
        
        Args:
            args: Command line arguments
            reference_file: Path to reference file
            mode: Pipeline mode ('standard', 'direct', or 'maf')
            
        Returns:
            tuple: (output_dir, temp_dir) - Paths to output and temporary directories
        """
        import logging
        import tempfile
        
        logger = logging.getLogger("ddPrimer")
        
        # Determine output directory
        if args.output:
            output_dir = args.output
        elif reference_file:
            if mode == 'maf' and args.maf_file:
                # For MAF files, use the parent directory of the "Alignments" folder
                maf_dir = os.path.dirname(os.path.abspath(args.maf_file))
                if os.path.basename(maf_dir) == "Alignments":
                    # If it's in an Alignments folder, go up one level
                    output_dir = os.path.join(os.path.dirname(maf_dir), "Primers")
                else:
                    # Otherwise use the parent directory of the MAF file
                    output_dir = os.path.join(os.path.dirname(maf_dir), "Primers")
            else:
                # Use the directory of the reference file
                input_dir = os.path.dirname(os.path.abspath(reference_file))
                output_dir = os.path.join(input_dir, "Primers")
        else:
            # Fallback to current directory if no reference file
            output_dir = os.path.join(os.getcwd(), "Primers")
        
        # Create output directory
        os.makedirs(output_dir, exist_ok=True)
        logger.debug(f"Created output directory: {output_dir}")
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp(prefix="ddprimer_temp_", dir=output_dir)
        logger.debug(f"Created temporary directory: {temp_dir}")
        
        return output_dir, temp_dir

    @staticmethod
    def cleanup_temp_directory(temp_dir):
        """
        Clean up temporary directory safely.
        
        Args:
            temp_dir: Path to temporary directory
            
        Returns:
            bool: True if cleanup was successful, False otherwise
        """
        import logging
        import shutil
        
        logger = logging.getLogger("ddPrimer")
        
        try:
            if temp_dir and os.path.exists(temp_dir):
                logger.debug(f"Cleaning up temporary directory: {temp_dir}")
                shutil.rmtree(temp_dir)
                return True
        except Exception as e:
            logger.warning(f"Error cleaning up temporary files: {e}")
            return False
        
        return True